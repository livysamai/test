import pandas as pd
import sys
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def dif_sheet(file1, file2, sheet_name, output_file):
    try:
        # 读取两个文件的指定 Sheet 页
        df1 = pd.read_excel(file1, sheet_name=sheet_name)
        df2 = pd.read_excel(file2, sheet_name=sheet_name)

        print(f"Loaded '{sheet_name}' from {file1} and {file2}")

        # 确保两个文件的行列数量一致
        if df1.shape != df2.shape:
            print("Error: Sheets have different dimensions.")
            return

        # 将整个 DataFrame 转换为字符串类型
        df1 = df1.applymap(str)
        df2 = df2.applymap(str)

        # 创建对比结果 DataFrame
        comparison_df = df1.copy()

        # 加载原始 Excel 文件
        wb1 = load_workbook(file1)
        ws1 = wb1[sheet_name]
        wb2 = load_workbook(file2)
        ws2 = wb2[sheet_name]

        # 定义填充颜色（黄色）
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for row in range(df1.shape[0]):
            for col in df1.columns:
                value1 = df1.at[row, col]
                value2 = df2.at[row, col]

                if value1 != value2:  # 比较数据
                    comparison_df.at[row, col] = f"{value1} -> {value2}"  # 标记差异

                    # 在 Excel 中标记差异为黄色
                    ws1.cell(row=row+2, column=col+1).fill = yellow_fill  # Excel 行列是从 1 开始的
                    ws2.cell(row=row+2, column=col+1).fill = yellow_fill  # 同样对第二个文件做标记

        print("Comparison completed. Saving results...")

        # 保存对比结果到新文件
        comparison_df.to_excel(output_file, index=False)

        # 保存标记后的 Excel 文件
        wb1.save("highlighted_" + file1)
        wb2.save("highlighted_" + file2)

        print(f"Comparison file saved to {output_file}")
        print(f"Highlighted files saved as highlighted_{file1} and highlighted_{file2}")

    except Exception as e:
        print(f"Error occurred: {e}")

if __name__ == "__main__":
    if len(sys.argv) != 5:
        print("Usage: python compare_sheets.py <file1> <file2> <sheet_name> <output_file>")
    else:
        file1 = sys.argv[1]
        file2 = sys.argv[2]
        sheet_name = sys.argv[3]
        output_file = sys.argv[4]
        dif_sheet(file1, file2, sheet_name, output_file)
