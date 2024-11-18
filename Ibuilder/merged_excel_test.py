import sys
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Fill, Alignment, Border
from openpyxl.utils import get_column_letter

def merge_files_to_sheets(input_files, output_file):
    """
    将多个 Excel 文件的内容分别写入一个新的 Excel 文件的不同 Sheet 页中，并保留格式
    """
    wb = Workbook()
    wb.remove(wb.active)  # 删除默认的 Sheet 页

    for file in input_files:
        try:
            # 读取 Excel 文件
            xls = pd.ExcelFile(file)
            for sheet_name in xls.sheet_names:
                # 读取 DataFrame
                df = pd.read_excel(file, sheet_name=sheet_name, header=None)
                sheet_title = f"{file.split('/')[-1].split('.')[0]}_{sheet_name}"[:31]  # 确保 Sheet 名称不超过 31 字符
                ws = wb.create_sheet(sheet_title)

                # 获取原始工作表的内容并复制到新的工作表中
                original_wb = load_workbook(file)
                original_ws = original_wb[sheet_name]

                # 复制列宽
                for col in range(1, original_ws.max_column + 1):
                    col_letter = get_column_letter(col)
                    ws.column_dimensions[col_letter].width = original_ws.column_dimensions[col_letter].width

                # 复制行高
                for row in range(1, original_ws.max_row + 1):
                    ws.row_dimensions[row].height = original_ws.row_dimensions[row].height

                # 复制每个单元格的内容和格式
                for row in original_ws.iter_rows(min_row=1, max_row=original_ws.max_row, min_col=1, max_col=original_ws.max_column):
                    for cell in row:
                        new_cell = ws.cell(row=cell.row, column=cell.column)
                        new_cell.value = cell.value

                        # 复制字体样式
                        if cell.font:
                            new_cell.font = Font(name=cell.font.name,
                                                 size=cell.font.size,
                                                 bold=cell.font.bold,
                                                 italic=cell.font.italic,
                                                 underline=cell.font.underline,
                                                 color=cell.font.color)

                        # 复制填充样式（背景颜色）
                        if cell.fill:
                            new_cell.fill = Fill(start_color=cell.fill.start_color,
                                                 end_color=cell.fill.end_color,
                                                 fill_type=cell.fill.fill_type)

                        # 复制对齐方式
                        if cell.alignment:
                            new_cell.alignment = Alignment(horizontal=cell.alignment.horizontal,
                                                           vertical=cell.alignment.vertical)

                        # 复制边框样式
                        if cell.border:
                            new_cell.border = Border(left=cell.border.left,
                                                      right=cell.border.right,
                                                      top=cell.border.top,
                                                      bottom=cell.border.bottom)

                        # 复制数字格式
                        new_cell.number_format = cell.number_format

                # 复制分页符（行和列的分页符）
                for row_break in original_ws.row_breaks:
                    ws.row_breaks.append(row_break)

                for col_break in original_ws.col_breaks:
                    ws.col_breaks.append(col_break)

        except Exception as e:
            print(f"Error processing file {file}: {e}", file=sys.stderr)
            continue

    # 保存合并后的文件
    try:
        wb.save(output_file)
        print(f"合并完成，保存到: {output_file}")
    except PermissionError:
        print(f"无法保存文件，请确保目标文件没有被占用或您有足够的写入权限。")

if __name__ == "__main__":
    # 从命令行接收文件路径
    input_files = sys.argv[1:-1]
    output_file = sys.argv[-1]
    if not input_files or not output_file:
        print("请提供输入文件和输出文件路径", file=sys.stderr)
        sys.exit(1)

    merge_files_to_sheets(input_files, output_file)
