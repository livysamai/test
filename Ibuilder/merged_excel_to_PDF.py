import sys
import pandas as pd
from openpyxl import load_workbook
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

def merge_excel_files(file_list, output_file):
    # 合并 Excel 文件
    merged_data = pd.DataFrame()
    for file in file_list:
        data = pd.read_excel(file)
        merged_data = pd.concat([merged_data, data], ignore_index=True)

    # 将合并的内容保存到 Excel 文件
    merged_data.to_excel(output_file, index=False)
    return output_file

def excel_to_pdf(excel_file, pdf_file):
    workbook = load_workbook(excel_file)
    sheet = workbook.active
    c = canvas.Canvas(pdf_file, pagesize=letter)
    width, height = letter  # 获取页面宽度和高度
    c.setFont("Helvetica", 8)  # 使用较小的字体以适应更多内容

    row_num = 0
    line_height = 12  # 设置行间距
    max_cols_per_page = 5  # 每页最多列数

    for row in sheet.iter_rows():
        col_num = 0
        for cell in row:
            x = 100 + col_num * 100  # 列的横向位置
            y = height - 100 - row_num * line_height  # 行的纵向位置
            c.drawString(x, y, str(cell.value))  # 绘制文本
            col_num += 1

        row_num += 1

        # 如果行数超过页面高度，则分页
        if row_num * line_height > height - 100:  # 如果内容超出页面
            c.showPage()  # 新的一页
            c.setFont("Helvetica", 8)  # 重新设置字体
            row_num = 0  # 重置行数

    c.save()

def main():
    # 从命令行参数中获取输入文件和输出路径
    input_files = sys.argv[1:-1]  # 输入文件列表
    output_file = sys.argv[-1]  # 输出文件路径

    # 合并 Excel 文件
    merged_file = merge_excel_files(input_files, output_file)

    # 转换为 PDF
    pdf_file = output_file.replace(".xlsx", ".pdf")
    excel_to_pdf(merged_file, pdf_file)

if __name__ == "__main__":
    main()
